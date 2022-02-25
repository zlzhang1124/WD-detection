#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright (c) 2022. Institute of Health and Medical Technology, Hefei Institutes of Physical Science, CAS
# @Time     : 2022/1/10 9:25
# @Author   : ZL.Z
# @Email    : zzl1124@mail.ustc.edu.cn
# @Reference: None
# @FileName : models.py
# @Software : Python3.6; PyCharm; Windows10 / Ubuntu 18.04.5 LTS (GNU/Linux 5.4.0-79-generic x86_64)
# @Hardware : Intel Core i7-4712MQ; NVIDIA GeForce 840M / 2*X640-G30(XEON 6258R 2.7G); 3*NVIDIA GeForce RTX3090
# @Version  : V2.0 - ZL.Z：2022/2/4
#             V1.0 - ZL.Z：2022/1/10 - 2022/1/12
#             First version.
# @License  : None
# @Brief    : 模型

from config import *
import datetime
import joblib
from matplotlib import pyplot as plt
from sklearn.base import clone
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split, GridSearchCV, StratifiedKFold, cross_val_score
from sklearn.calibration import CalibratedClassifierCV
from sklearn.pipeline import Pipeline
from sklearn.svm import SVC
from sklearn.metrics import roc_curve, make_scorer, confusion_matrix, plot_confusion_matrix, \
	roc_auc_score, accuracy_score, precision_recall_fscore_support
from openpyxl import load_workbook


class ModelBase:
	"""模型基类"""

	def __init__(self, data_dir: str = "", method: str = "DMFCC-CEEMDAN",
	             model_name: str = "", shuffle=True, random_state=rs):
		"""
		初始化
		:param data_dir: 数据集文件路径
		:param method: 特征方法，包括MFCC与(DMFCC/HCC)-(VMD/CEEMDAN/EMD)
		:param model_name: 模型名称
		:param shuffle: 是否打乱数据集，默认打乱
		:param random_state: 随机种子
		"""
		data_dir = os.path.normpath(data_dir)
		feat_data = pd.read_pickle(os.path.join(data_dir, 'features_' + method.split('-')[0] + '.pkl'))
		_x_data, y_label = feat_data[method], feat_data.iloc[:, 1:2]  # 排除性别及年龄特征
		x_data = []
		for i_subj in _x_data:
			x_data.append(i_subj)
		x_data = np.array(x_data)  # shape=[样本数，特征维数]
		self.rs = random_state
		train_data, test_data, train_label, test_label = train_test_split(x_data, y_label, random_state=self.rs,
		                                                                  test_size=0.3, shuffle=shuffle,
		                                                                  stratify=y_label)
		self.train_label, self.test_label = np.array(train_label).ravel(), np.array(test_label).ravel()
		# 划分数据后再进行数据处理，避免测试集数据泄露
		ss = StandardScaler()  # 标准化特征
		pipe = Pipeline([('ss', ss)])
		self.train_data = pipe.fit_transform(train_data)
		self.test_data = pipe.transform(test_data)
		self.data_all = np.vstack((self.train_data, self.test_data))  # 拼接训练集和测试集
		self.label_all = np.vstack((self.train_label.reshape((-1, 1)), self.test_label.reshape((-1, 1)))).ravel()
		self.pipe_save_dir = os.path.join('./models', 'pipe')
		self.model_save_dir = os.path.join('./models', 'clf')  # 保存模型路径
		self.model_save_dir_final = os.path.join('./models', 'final')  # 最终训练全部数据保存模型路径
		self.fig_save_dir = os.path.join('./results', 'clf/')  # 保存结果曲线图片路径
		if not os.path.exists(self.pipe_save_dir):
			os.makedirs(self.pipe_save_dir)
		if not os.path.exists(self.model_save_dir):
			os.makedirs(self.model_save_dir)
		if not os.path.exists(self.model_save_dir_final):
			os.makedirs(self.model_save_dir_final)
		if not os.path.exists(self.fig_save_dir):
			os.makedirs(self.fig_save_dir)
		joblib.dump(pipe, os.path.join(self.pipe_save_dir, 'pipeline_processed.m'))  # 保存数据处理模型，后续新数据transform使用
		self.model_name = model_name
		self.method = method
		self.model_file = os.path.join(self.model_save_dir, f'{self.model_name}_{self.method}.m')
		self.model_file_final = os.path.join(self.model_save_dir_final,
		                                     f'clf/{self.model_name}_{self.method}_final.m')
		self.fig_file = os.path.join(self.fig_save_dir, f"ROC_{self.model_name}_{self.method}.png")
		self.wb = load_workbook(perform_comp_f)  # 模型性能比较的EXCEL文件
		self.sheet = self.wb['Sheet1']

	def model_train(self):
		"""
		模型训练
		"""
		pass

	def model_evaluate(self, fig=False, sen_spe=False):
		"""
		模型评估
		:param fig: 显示测试集的ROC曲线以及预测曲线
		:param sen_spe: ROC曲线标题中是否显示敏感性和特异性值
		:return: 评估测试集的准确率acc, F1值f1_score, ROC曲线下面积roc_auc,
				 敏感性sensitivity（即召回率recall）, 特异性specificity
		"""
		if os.path.exists(self.model_file):  # 存在已训练模型且设置加载，
			print("----------加载分类模型：{}----------".format(self.model_file))
			clf_model = joblib.load(self.model_file)  # 加载已训练模型
		else:
			print("分类模型不存在，无法评估，请先训练")
			return None
		y_preds = clf_model.predict(self.test_data)
		y_pred_proba = clf_model.predict_proba(self.test_data)
		acc = accuracy_score(self.test_label, y_preds)
		roc_auc = roc_auc_score(self.test_label, y_pred_proba[:, 1])
		precision, recall, f1_score, support = precision_recall_fscore_support(self.test_label, y_preds,
		                                                                       average='binary')  # 测试集各项指标
		sen = recall
		spec = scorer_sensitivity_specificity(self.test_label, y_preds)
		print('Test set Accuracy: %.4f\nTest set F1 score: %.4f\nTest set ROC-AUC: %.4f'
		      '\nTest set Sensitivity (Recall): %.4f\nTest set Specificity: %.4f\n' %
		      (acc, f1_score, roc_auc, sen, spec))
		index = 0
		for column_cell in self.sheet.iter_rows():  # 遍历行
			index += 1
			if column_cell[0].value == self.method:  # 每行的首个为对应的Features
				break
		self.sheet['G' + str(index)] = round(acc, 4)
		self.sheet['H' + str(index)] = round(f1_score, 4)
		self.sheet['I' + str(index)] = round(sen, 4)
		self.sheet['J' + str(index)] = round(spec, 4)
		self.sheet['K' + str(index)] = round(roc_auc, 4)
		self.wb.save(perform_comp_f)  # 将结果写入模型性能比较的EXCEL文件
		if fig:  # 显示测试集的ROC曲线与混淆矩阵
			if sen_spe:
				title_t = f'ROC Curve of Detecting WD Using {self.method} ({self.model_name})\n' \
				          f'Sensitivity = {sen:.2f}, Specificity = {spec:.2f}'
			else:
				title_t = f'ROC Curve of Detecting WD Using {self.method} ({self.model_name})'
			plt.figure(figsize=(8, 6))
			plt.title(title_t, fontdict={'family': font_family, 'size': 16})
			plt.xlabel('False Positive Rate', fontdict={'family': font_family, 'size': 14})
			plt.ylabel('True Positive Rate', fontdict={'family': font_family, 'size': 14})
			plt.plot([0, 1], [0, 1], c='gray', lw=1.2, ls='--')
			fpr, tpr, thresholds = roc_curve(self.test_label, y_pred_proba[:, 1])
			plt.plot(fpr, tpr, label=f'ROC curve of {self.method} (area = {roc_auc:.2f})', c='r', ls='-', lw=2)
			plt.legend(loc="lower right", prop={'family': font_family, 'size': 12})
			plt.xlim(0.0, 1.0)
			plt.ylim(0.0, 1.0)
			for sp in plt.gca().spines:
				plt.gca().spines[sp].set_color('black')
				plt.gca().spines[sp].set_linewidth(1)
			plt.gca().tick_params(direction='out', color='black', length=5, width=1)
			plt.grid(False)
			plt.tight_layout()
			if not os.path.exists(os.path.dirname(self.fig_file)):
				os.makedirs(os.path.dirname(self.fig_file))
			plt.savefig(self.fig_file, dpi=600)
			plt.savefig(self.fig_file.replace('.png', '.svg'), format='svg')
			# plt.show()
			plt.close('all')
			# 绘制混淆矩阵
			plt.figure(figsize=(8, 6))
			disp = plot_confusion_matrix(clf_model, self.test_data, self.test_label, normalize='true',
			                             colorbar=False, cmap='Blues', values_format='.2f')
			disp.ax_.set_title(f'Confusion Matrix of Detecting WD Using {self.method} ({self.model_name})',
			                   fontdict={'family': font_family, 'size': 16})
			disp.ax_.set_xlabel('Predicted Class', fontdict={'family': font_family, 'size': 14})
			disp.ax_.set_ylabel('True Class', fontdict={'family': font_family, 'size': 14})
			for sp in plt.gca().spines:
				plt.gca().spines[sp].set_color('black')
				plt.gca().spines[sp].set_linewidth(1)
			plt.gca().tick_params(direction='out', color='black', length=5, width=1)
			plt.grid(False)
			plt.tight_layout()
			save_fig_path = self.fig_file.replace(f'ROC_', 'confusion_mat_')
			plt.savefig(save_fig_path, dpi=600)
			plt.savefig(save_fig_path.replace('.png', '.svg'), format='svg')
			# plt.show()
			plt.close('all')
		return acc, f1_score, roc_auc, sen, spec


class SvmModel(ModelBase):
	"""SVM模型"""

	def __init__(self, model_name='SVM', **kwargs):
		"""
		初始化
		:param **kwargs: ModelBase类__init__参数
		"""
		super(SvmModel, self).__init__(model_name=model_name, **kwargs)

	def model_train(self):
		"""
		模型训练
		:return: 分类模型
		"""
		# rs=323,cv=0.8236±0.115,test=0.8611(acc, DMFCC-CEEMDAN): {'C': 10, 'gamma': 'auto'}
		# 网格搜索+检查验证超参数
		param_grid = {'base_estimator__C': [0.01, 0.1, 1, 10, 100],
		              'base_estimator__gamma': [0.005, 0.01, 0.05, 0.1, 1, 10, 'scale', 'auto']}
		skf = StratifiedKFold(n_splits=10, shuffle=True, random_state=self.rs)  # 10折交叉验证，分层采样
		model_svc = SVC(probability=True, random_state=self.rs)
		# 使用CalibratedClassifierCV概率校准估计器SVC，避免后期predict和predict_proba结果不一致
		model_cal = CalibratedClassifierCV(model_svc, cv=skf, n_jobs=1)  # 仅校准最优参数对应的模型
		model_grid = GridSearchCV(model_cal, param_grid, cv=skf, scoring='accuracy', n_jobs=-1)
		model_grid.fit(self.train_data, self.train_label)
		model_clf = model_grid.best_estimator_  # 最优模型
		joblib.dump(model_clf, self.model_file)  # 保存模型
		model_final = clone(model_clf)  # 深复制不会改变
		model_final.fit(self.data_all, self.label_all)  # 对包括训练集和测试集的全部数据进行最终fit
		if not os.path.exists(os.path.dirname(self.model_file_final)):
			os.makedirs(os.path.dirname(self.model_file_final))
		joblib.dump(model_final, self.model_file_final)  # 保存最终模型
		print("------ The best Classification params -------")
		print(model_grid.best_params_)
		# 输出最优参数的10折交叉验证的各项指标
		acc = get_cv_res(model_clf, self.train_data, self.train_label, "accuracy", skf)
		print(f'CV Accuracy: {acc[0]}±{acc[1]}')
		f1 = get_cv_res(model_clf, self.train_data, self.train_label, "f1", skf)
		print(f'CV F1 score: {f1[0]}±{f1[1]}')
		roc_auc = get_cv_res(model_clf, self.train_data, self.train_label, "roc_auc", skf)
		print(f'CV ROC-AUC: {roc_auc[0]}±{roc_auc[1]}')
		sen = get_cv_res(model_clf, self.train_data, self.train_label, "recall", skf)
		print(f'CV Sensitivity (Recall): {sen[0]}±{sen[1]}')
		spec = get_cv_res(model_clf, self.train_data, self.train_label,
		                  make_scorer(scorer_sensitivity_specificity), skf)
		print(f'CV Specificity: {spec[0]}±{spec[1]}')
		index = 0
		for column_cell in self.sheet.iter_rows():  # 遍历行
			index += 1
			if column_cell[0].value == self.method:  # 每行的首个为对应的Features
				break
		self.sheet['B' + str(index)] = f'{acc[0]}±{acc[1]}'
		self.sheet['C' + str(index)] = f'{f1[0]}±{f1[1]}'
		self.sheet['D' + str(index)] = f'{sen[0]}±{sen[1]}'
		self.sheet['E' + str(index)] = f'{spec[0]}±{spec[1]}'
		self.sheet['F' + str(index)] = f'{roc_auc[0]}±{roc_auc[1]}'
		self.sheet['L' + str(index)] = str(model_grid.best_params_)
		self.wb.save(perform_comp_f)  # 将结果写入模型性能比较的EXCEL文件
		return model_clf


def scorer_sensitivity_specificity(y_true, y_pred, sen_spec=False):
	"""
	敏感性特异性指标
	:param y_true: 真实值
	:param y_pred: 预测概率
	:param sen_spec: True返回特异性，否则返回敏感性
	:return: 敏感性、特异性
	"""
	tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()
	sen = tp / (tp + fn)
	spec = tn / (tn + fp)
	if sen_spec:
		return sen
	else:
		return spec


def get_cv_res(model, features, labels, score, cv=None, n_jobs=-1):
	"""
    交叉验证并获取评价指标的均值、标准差
    :param model: 模型：最优参数对应的分类或回归
    :param features: 特征
    :param labels: 标签
    :param score: 评价指标
    :param cv: 交叉验证拆分策略，默认10折交叉验证
    :param n_jobs: 并行运行的作业数
    :return: 评价指标的均值、标准差
    """
	if cv is None:
		cv = 10
	res = cross_val_score(model, features, labels, cv=cv, scoring=score, n_jobs=n_jobs)
	return round(res.mean(), 4), round(res.std(), 4)


def dmfcc_roc_plot():
	"""
	绘制在不同信号分解方法下的DMFCC特征，在测试集上对应的ROC曲线
	:return: None
	"""
	method_l = ['CEEMDAN', 'VMD', 'EMD']
	clf_model = {'CEEMDAN': '', 'VMD': '', 'EMD': ''}
	for i_md in method_l:
		model_file = os.path.join(current_path, f'models/clf/SVM_DMFCC-{i_md}.m')
		if os.path.exists(model_file):  # 存在已训练模型且设置加载，
			print("----------加载分类模型：{}----------".format(model_file))
			clf_model[i_md] = joblib.load(model_file)  # 加载已训练模型
		else:
			print("分类模型不存在，请先训练")
			return None
	plt.figure(figsize=(8, 6))
	plt.title('ROC curves obtained with DMFCC using different signal decomposition methods',
	          fontdict={'family': font_family, 'size': 16})
	plt.xlabel('False Positive Rate', fontdict={'family': font_family, 'size': 14})
	plt.ylabel('True Positive Rate', fontdict={'family': font_family, 'size': 14})
	plt.plot([0, 1], [0, 1], c='gray', lw=1.2, ls='--')
	lcs = {'CEEMDAN': ['r', '-'], 'VMD': ['b', ':'], 'EMD': ['g', '--']}
	for i_md in method_l:
		__model = SvmModel(data_dir=feat_path, method='DMFCC-' + i_md, model_name='SVM')
		y_pred_proba = clf_model[i_md].predict_proba(__model.test_data)
		roc_auc = roc_auc_score(__model.test_label, y_pred_proba[:, 1])
		fpr, tpr, thresholds = roc_curve(__model.test_label, y_pred_proba[:, 1])
		plt.plot(fpr, tpr, label=f'ROC curve of {i_md} (area = {roc_auc:.2f})', c=lcs[i_md][0],
		         ls=lcs[i_md][1], lw=2)
	plt.legend(loc="lower right", prop={'family': font_family, 'size': 12})
	plt.xlim(0.0, 1.0)
	plt.ylim(0.0, 1.0)
	for sp in plt.gca().spines:
		plt.gca().spines[sp].set_color('black')
		plt.gca().spines[sp].set_linewidth(1)
	plt.gca().tick_params(direction='out', color='black', length=5, width=1)
	plt.grid(False)
	plt.tight_layout()
	fig_file = os.path.join(res_path, 'ROC_DMFCC.png')
	if not os.path.exists(os.path.dirname(fig_file)):
		os.makedirs(os.path.dirname(fig_file))
	plt.savefig(fig_file, dpi=600)
	plt.savefig(fig_file.replace('.png', '.svg'), format='svg')
	plt.show()
	plt.close('all')


if __name__ == "__main__":
	start_time = datetime.datetime.now()
	print(f"---------- Start Time ({os.path.basename(__file__)}): {start_time.strftime('%Y-%m-%d %H:%M:%S')} ----------")
	current_path = os.getcwd()
	feat_path = os.path.join(current_path, "data/features")
	res_path = os.path.join(current_path, r"results")
	perform_comp_f = os.path.join(current_path, r"results/PerformanceComparison.xlsx")

	for md in ["HCC-VMD", "HCC-CEEMDAN", "HCC-EMD", "DMFCC-VMD", "DMFCC-CEEMDAN", "DMFCC-EMD", "MFCC", ]:
		print(f"------- {md} features -------\n")
		_model = SvmModel(data_dir=feat_path, method=md, model_name='SVM')
		_model.model_train()
		_model.model_evaluate(fig=True, sen_spe=True)
	dmfcc_roc_plot()

	end_time = datetime.datetime.now()
	print(f"---------- End Time ({os.path.basename(__file__)}): {end_time.strftime('%Y-%m-%d %H:%M:%S')} ----------")
	print(f"---------- Time Used ({os.path.basename(__file__)}): {end_time - start_time} ----------")
	with open(r"./results/finished.txt", "w") as ff:
		ff.write(f"------------------ Started at {start_time.strftime('%Y-%m-%d %H:%M:%S')} "
				 f"({os.path.basename(__file__)}) -------------------\r\n")
		ff.write(f"------------------ Finished at {end_time.strftime('%Y-%m-%d %H:%M:%S')} "
				 f"({os.path.basename(__file__)}) -------------------\r\n")
		ff.write(f"------------------ Time Used {end_time - start_time} "
				 f"({os.path.basename(__file__)}) -------------------\r\n")
